import csv
import pandas as pd
from openpyxl import workbook, load_workbook

try:
    with open(r'C:\Users\WKYTUK0\Downloads\data (1).csv') as csvFile:
        file_reader = csv.reader(csvFile, delimiter=',',quotechar = '|')
        df = list(file_reader)

    #for i in range(len(df)):
     #   print(df[i])
except IOError:
    print(IOError)

#Give how many columns are in csv file
alpha = 0
for i in range(len(df[0])):
    alpha = alpha + 1

beta = 0
for j in range(len(df)):
    beta = beta+1


#Gives how many rows are in csv file
#print(len(df[0]))

#filename = r"C:\Users\WKYTUK0\OneDrive-Deere&Co\OneDrive - Deere & Co\Documents\EDP Rotation 1\Excel Documents\Operator_DataLog - Copy.xlsx"
filename = r"C:\Users\WKYTUK0\Downloads\Python.xlsx"
wb = load_workbook(filename)

ws = wb.active

j=0
for x in range(beta - 5):
    for y in range(len(df[1])):

        if x > 0 and y == 2:
            gamma = df[x][2] + df[x][3]
            ws.cell(row=x+1, column=y+1, value=f'{gamma}')

        elif x > 0 and y == 1:
            ws.cell(row=x + 1, column=y + 1, value=f'{df[x][y]}')

        elif x > 0 and y == 3:
            continue

        #elif x > 0 and y == 7:
         #   ws.cell(row=x + 1, column=y, value=f'{df[x][y]}')
            #ws.cell(row=x + 1, column=y+1, value=f'{df[x][8]}')

        elif x > 0 and y > 3:
            try:
                ws.cell(row= x + 1, column=y, value=f'{df[x][y]}')
            except:
                print(f'Line {x}')

        elif x == 0:
            if y == 8:
                pass
            else:
                ws.cell(row=x + 1, column=y+1, value=f'{df[x][y]}')

        else:
                ws.cell(row=x + 1, column=y+1, value=f'{df[x][y]}')

wb.save(filename)
print('Program completed')

#print(ws['A1'].value)



